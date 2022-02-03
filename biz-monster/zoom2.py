import webbrowser as web
import pyautogui
from time import sleep
import pyperclip
# from selenium import webdriver
# from selenium.webdriver.common.keys import Keys
# from bs4 import BeautifulSoup
import csv
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\jayka\PycharmProjects\pythonProject\test\school.xlsx')
sheet = wb.active

wb2 = load_workbook(r'C:\Users\jayka\PycharmProjects\pythonProject\test\nodata.xlsx') #nodata
sheet2 = wb2.active
max_row = sheet2.max_row

r = 2 #row number for write a data
rows = []
with open(rb"C:\Users\jayka\PycharmProjects\pythonProject\test\links2.csv") as csvfile:  # reading csv file
    csvreader = csv.reader(csvfile)  # creating a csv reader object

    fields = next(csvreader)  # extracting field names through first row"

    for row in csvreader:  # extracting each data row one by one
        rows.append(row)

    lines = csvreader.line_num  # get total number of rows
    print("Total no. of rows: %d" % (lines))

    rnum = 0
    for school in rows[:lines]:  # lines #school[0]=name school[1]=school url school[2]=zoominfo url
        original_link = school[1].replace("https://", '').replace('http://', '').split('/')[0]
        school_link = school[2]
        webpage = web.open(school_link.replace('/pic/', '/c/'), 1)
        # if rnum == 0:
        #     sleep(30)
        #     rnum += 1
        sleep(6)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'c')

        lpage = str(pyperclip.paste()).split('\r')  # whole page text

        new_lpage = []
        for element in lpage:  # remove new lines and space from the list
            new_lpage.append(element.strip())

        ltest = list(filter(lambda x: x, new_lpage))  # remove empty list from main list
        print("find website:")
        print(ltest)
        ind = 0  # index of list(test)
        for flink in ltest:
            if "website:" in flink.lower():
                sch_link = ltest[ind + 1]
                break
            ind += 1
        try:
            if original_link.replace('www.','') == ltest[ind + 1].replace('www.',''): ####### if link is match then start the next programm
                print(ltest[ind + 1])
                print("original link:"+ original_link)
                first_page = 0

                for pnum in range(4):
                    try:
                        webpage = web.open(school[2].replace('/c/','/pic/') + '?pageNum={}'.format(pnum + 1), 3)
                        print(webpage)
                        sleep(7)
                        pyautogui.hotkey('ctrl', 'a')
                        pyautogui.hotkey('ctrl', 'c')

                        page = str(pyperclip.paste()).replace('emailEmail','').replace('directDirect','').split('\r') #whole page text
                        new_list = []
                        for element in page:    #remove new lines and space from the list
                            new_list.append(element.strip())

                        test = list(filter(lambda x: x, new_list))  # remove empty list from main list
                        print(test)

                        if "404" in test:
                            break

                        lst = ["principal", 'education' , 'administrative', 'head of school', 'associate', 'executive', 'department head',
                               'outreach', 'finance', 'financial', 'admissions', 'admission', 'counseling', 'communications',
                               'communication', 'dean', 'enrollment', 'marketing', 'president', 'vice president']

                        n = 0 #list which has a data

                        for word in test:
                            if "Profile Picture" in word:
                                res = [ele for ele in lst if (ele in test[n + 2].lower())]
                                if bool(res):
                                    print(test[n + 1])  # name
                                    sheet.cell(row=r, column=1).value = test[n + 1] # name
                                    print(test[n + 2])  # job
                                    sheet.cell(row=r, column=2).value = test[n + 2] # job
                                    sheet.cell(row=r, column=3).value = school[0]  #school[0]=name school[1]=school url school[2]=zoominfo url
                                    sheet.cell(row=r, column=4).value = school[1] #school[0]=name school[1]=school url school[2]=zoominfo url
                                    sheet.cell(row=r, column=5).value = ltest[ind + 1] #link which is avalable in zoominfo

                                    r += 1  # row number increse for write a data
                                else:
                                    pass

                            elif "profile photo" in word:
                                res = [ele for ele in lst if (ele in test[n + 2].lower())]
                                if bool(res):
                                    print(test[n + 1])  # name
                                    sheet.cell(row=r, column=1).value = test[n + 1] # name
                                    print(test[n + 2])  # job
                                    sheet.cell(row=r, column=2).value = test[n + 2] # job
                                    sheet.cell(row=r, column=3).value = school[0] #school[0]=s name school[1]=school url school[2]=zoominfo url
                                    sheet.cell(row=r, column=4).value = school[1] #school[0]=s name school[1]=school url school[2]=zoominfo url
                                    sheet.cell(row=r, column=5).value = ltest[ind + 1] #link which is avalable in zoominfo

                                    r += 1  # row number increse for write a data
                                else:
                                    pass

                            n += 1 #find next list which has a data

                        if first_page > 0:
                            pyautogui.hotkey('ctrl', 'w')  # close the tab
                        first_page += 1
                        print(first_page)
                        wb.save(r'C:\Users\jayka\PycharmProjects\pythonProject\test\school.xlsx')
                    except: #page 1,2,3
                        break
            else:
                sheet2['A1'] = 'School name'
                sheet2['B1'] = 'school url'
                sheet2.cell(row=max_row + 1, column=1).value = school[0]
                sheet2.cell(row=max_row + 1, column=2).value = school[1]
            wb2.save(r'C:\Users\jayka\PycharmProjects\pythonProject\test\nodata.xlsx')
        except: # if original_link == ltest[ind + 1]: checking page is real or not
            pass



#links.csv ----- school.xlsx



