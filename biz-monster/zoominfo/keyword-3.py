import webbrowser as web
import pyautogui
import pyperclip
from time import sleep
import os
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\jayka\PycharmProjects\pythonProject\nces\keyword.xlsx')
sheet = wb.active
max_row = sheet.max_row
print(max_row)

web.open('https://mail.google.com/mail/u/0/#inbox?compose=new')
sleep(12)
n=2 #rows number in excel

lines = max_row+1

for i in range(1,lines):  # lines
    print(str(sheet.cell(row=i+1, column=1).value),str(sheet.cell(row=i+1, column=5).value)) #i[0] = name i[5] = url
    sleep(0)
    pyautogui.click(x=870, y=310, clicks=1, interval=2, button='left') #Point(x=1044, y=412) Click on mailbox (To)
    domain = str("@" + sheet.cell(row=i+1, column=5).value).replace('www.','').replace('http://','').replace('https://','').split('/')[0]
    pyautogui.write(str(sheet.cell(row=i+1, column=1).value) + " " + domain)
    sleep(7)
    pyautogui.click(x=869, y=336, clicks=1, interval=2, button='left') #Click on email which is found
    sleep(0)
    print(pyautogui.position())
    pyautogui.click(x=870, y=310, clicks=1, interval=2, button='left') #Click on mail box again for copy thon mail
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('ctrl', 'x')
    email = str(pyperclip.paste())
    # just checking we got what we need
    print(email)
    try:
        if " " in email:
            sheet.cell(row=n, column=6).value = None
            n += 1
        else:
            sheet.cell(row=n , column=6).value = email
            n += 1
    except:
        print("Oh no. We didn't highlight an str value.")
        n += 1

    wb.save(r'C:\Users\jayka\PycharmProjects\pythonProject\nces\keyword.xlsx')

#close tab and shutdown PC
pyautogui.hotkey('alt', 'F4')
os.system("shutdown /s /t 1")
#book3.csv --> school.xlsx