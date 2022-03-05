import webbrowser as web
import pyautogui
from time import sleep
import pyperclip
# import csv
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\jayka\PycharmProjects\pythonProject\nces\keyword.xlsx')
sheet = wb.active
r = 1 + sheet.max_row


# r = 2 #row number for write a data

#admission #marketing #recruitment #enrollment
u = "https://www.zoominfo.com/people-search/location-usa--texas-industry-colleges-universities-title-"
# state = "NJ"

keylst = ['admission', 'marketing','recruitment', 'enrollment' ] #, 'marketing', 'recruitment'
for keyw in keylst:
    for pnum in range(5):
        try:
            webpage = web.open(u + "{}".format(keyw) + '?pageNum={}'.format(pnum+1))
            print(webpage)

            sleep(5)
            pyautogui.hotkey('ctrl', 'a')
            sleep(1)
            pyautogui.keyDown('ctrl')
            pyautogui.press('c')
            pyautogui.keyUp('ctrl')
            # pyautogui.hotkey('ctrl', 'c')

            page = str(pyperclip.paste()).replace('email','').replace('Email','').replace('phone','').replace('Direct','').split('\r') #whole page text
            new_list = []
            for element in page:    #remove new lines and space from the list
                new_list.append(element.strip())

            test = list(filter(lambda x: x, new_list))  # remove empty list from main list
            # print(test)

            for cloud in test:
                if "browser" in cloud:
                    sleep(3)
                    pyautogui.hotkey('ctrl', 'a')
                    sleep(1)
                    pyautogui.keyDown('ctrl')
                    pyautogui.press('c')
                    pyautogui.keyUp('ctrl')
                    # pyautogui.hotkey('ctrl', 'c')

                    page = str(pyperclip.paste()).replace('email', '').replace('Email', '').replace('phone', '').replace(
                        'Direct', '').split('\r')  # whole page text
                    new_list = []
                    for element in page:  # remove new lines and space from the list
                        new_list.append(element.strip())

                    test = list(filter(lambda x: x, new_list))  # remove empty list from main list

            cou = []
            for count in test:
                if "United States" == count:
                   cou.append("United States")
            print("len of test" + str(len(cou)))

            if len(cou) == 0:
                break
            lst = ["principal" , 'head of school', 'recruitment',
                    'admissions', 'admission', 'provost', 'admission director', 'enrollment director', 'dean of student', 'associate dean', 'enrollment', 'marketing', 'president']

            n = 0 #list which has a data

            for word in test:
                if "United States" in word:
                    res = [ele for ele in lst if (ele in test[n - 1].lower())]
                    if bool(res):
                        print(test[n - 2])  # name
                        sheet.cell(row=r, column=1).value = test[n - 2] # name
                        print(test[n - 1])  # job
                        if test[n - 1].strip().startswith('or'):
                            sheet.cell(row=r, column=2).value = test[n - 1].replace('or','Director') # job
                        else:
                            sheet.cell(row=r, column=2).value = test[n - 1].replace(' or ',' Director').replace(' or, ',' Director') # job
                        try:
                            print(test[n] + ", " + test[n + 2] + ", " + test[n + 4])  # location
                            sheet.cell(row=r, column=3).value = test[n] + ", " + test[n + 2] + ", " + test[n + 4] #clg location
                        except:
                            print(test[n] + ", " + test[n + 2])  # location
                            sheet.cell(row=r, column=3).value = test[n] + ", " + test[n + 2] # clg location

                        try:
                            if test[n-2] == test[n+5]:
                                print(test[n+6]) #clg name
                                sheet.cell(row=r, column=4).value = test[n+6]  # clg name
                            elif test[n-2] == test[n+3]:
                                print(test[n+4]) #clg name
                                sheet.cell(row=r, column=4).value = test[n+4]  # clg name
                            elif test[n-2] == test[n+1]:
                                print(test[n+2]) #clg name
                                sheet.cell(row=r, column=4).value = test[n+2] # clg name
                        except:
                            pass

                        r += 1  # row number increse for write a data



                n += 1 #find next list which has a data



        except: #page 1,2,3
            pass
        pyautogui.hotkey('ctrl', 'w')

wb.save(r'C:\Users\jayka\PycharmProjects\pythonProject\nces\keyword.xlsx')
clgdict = {}
wb1 = load_workbook(r'C:\Users\jayka\PycharmProjects\pythonProject\nces\clg-2.xlsx')
sheet1 = wb1.active
max_row1 = sheet1.max_row
print(max_row1)

for r in range(1,max_row1 + 1):
    # if str(sheet1.cell(row=r, column=3).value) == state:
    clgdict[str(sheet1.cell(row=r, column=1).value)] = str(sheet1.cell(row=r, column=2).value)

print("dictionary: " + str(clgdict))

wb2 = load_workbook(r'C:\Users\jayka\PycharmProjects\pythonProject\nces\keyword.xlsx')
sheet2 = wb2.active
max_row2 = sheet2.max_row

clst = []
for fclg in range(1,max_row2+1): #keyword xsl
    clst.append(sheet2.cell(row=fclg, column=4).value)
rn = 1
for cname in clst:
    try:
        print(clgdict[cname])
        sheet2.cell(row=rn, column=5).value = clgdict[cname]
        rn += 1
    except:
        rn += 1

wb2.save(r'C:\Users\jayka\PycharmProjects\pythonProject\nces\keyword.xlsx')
#links-2.csv ----- College.xlsx



