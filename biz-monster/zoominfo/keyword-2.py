from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from time import sleep
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\jayka\PycharmProjects\pythonProject\nces\testing.xlsx')
sheet = wb.active
# sheet['A1'] = 'College Name'
# sheet['B1'] = 'College URL'
max_row = sheet.max_row
print(max_row)

ser = Service("D:\chromedriver")
op = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=ser, options=op)

lines = max_row+1


for c in range(460,490): # 0 to 10 ## 10 to 20
    url = 'https://google.com/search?q=' + "{0}".format(str(sheet.cell(row=c + 1, column=4).value)).replace('&','%26').replace(' ','+')
    driver.get(url)
    page = driver.page_source
    soup = BeautifulSoup(page, 'html.parser')
    # print(soup.prettify())

    div = soup.find_all('div', {"class": "yuRUbf"})
    tag = div[0].find('a', href=True)
    # print(tag)
    link = tag['href'].replace('/url?q=', '').split('&', 1)[0]
    print(link)

    sheet.cell(row=c + 1, column=5).value = link #link of clg

wb.save(r'C:\Users\jayka\PycharmProjects\pythonProject\nces\testing.xlsx')
driver.quit()