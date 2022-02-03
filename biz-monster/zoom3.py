import webbrowser as web
import pyautogui
import pyperclip
from time import sleep
import csv
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\jayka\PycharmProjects\pythonProject\test\school.xlsx')

sheet = wb.active

web.open('https://mail.google.com/mail/u/0/#inbox?compose=new')

n=2 #rows number in excel
rows = []
with open(rb"C:\Users\jayka\PycharmProjects\pythonProject\test\book3.csv") as csvfile:  # reading csv file
    csvreader = csv.reader(csvfile)  # creating a csv reader object

    fields = next(csvreader)  # extracting field names through first row"

    for row in csvreader:  # extracting each data row one by one
        rows.append(row)

    lines = csvreader.line_num  # get total number of rows
    print("Total no. of rows: %d" % (lines))

    for i in rows[n-2:lines]:  # lines
        print(i[0],i[3]) #i[0] = name i[3] = url
        sleep(1)
        pyautogui.click(x=870, y=310, clicks=1, interval=2, button='left') #Point(x=1044, y=412) Click on mailbox (To)
        domain = i[3].replace('www.','').replace('http://','@').replace('https://','@').split('/')[0]
        pyautogui.write(i[0] + " " + domain)
        sleep(7)
        pyautogui.click(x=869, y=336, clicks=1, interval=2, button='left') #Click on email which is found
        sleep(0.5)
        print(pyautogui.position())
        pyautogui.click(x=870, y=310, clicks=1, interval=2, button='left') #Click on mail box again for copy thon mail
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.hotkey('ctrl', 'x')
        email = str(pyperclip.paste())
        # just checking we got what we need
        print(email)
        try:
            if " " in email:
                sheet.cell(row=n, column=6).value = None
                n += 1
            else:
                sheet.cell(row=n , column=6).value = email
                n += 1
        except:
            print("Oh no. We didn't highlight an str value.")
            n += 1

        wb.save(r'C:\Users\jayka\PycharmProjects\pythonProject\test\school.xlsx')


#book3.csv --> school.xlsx





















# from selenium import webdriver
# from time import sleep
#
# chop = webdriver.ChromeOptions()
# chop.add_extension('email.crx')
# driver = webdriver.Chrome("D:\chromedriver", chrome_options=chop)
# driver.maximize_window()
#
# driver.get("https://www.google.com/gmail/")
# driver.switch_to.window(driver.window_handles[0])
# driver.close()
#
# email = driver.find_element_by_id('identifierId')
# email.send_keys("jp738317@gmail.com")
# btn = driver.find_element_by_xpath('//*[@id="identifierNext"]/div/button')




