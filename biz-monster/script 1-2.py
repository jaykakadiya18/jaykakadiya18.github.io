import csv
from openpyxl import load_workbook

wb2 = load_workbook(r'C:\Users\jayka\PycharmProjects\pythonProject\test\nodata.xlsx')  # nodata
sheet2 = wb2.active
max_row = sheet2.max_row
print("nodata rows"+str(max_row))
rows = []
with open(rb"C:\Users\jayka\PycharmProjects\pythonProject\test\links.csv") as csvfile:  # reading csv file
    csvreader = csv.reader(csvfile)  # creating a csv reader object

    fields = next(csvreader)  # extracting field names through first row"

    for row in csvreader:  # extracting each data row one by one
        rows.append(row)

    lines = csvreader.line_num  # get total number of rows
    print("Total no. of rows: %d" % (lines))

    f = open('links2.csv', 'w')  # title line
    f.write("School Name,School url, Info Link\n")  # state name in csv

    for i in rows[:lines]:  # lines
        if i[2] == 'none':
            sheet2.cell(row=max_row + 1, column=1).value = i[0]
            sheet2.cell(row=max_row + 1, column=2).value = i[1]
            max_row +=1
            print(max_row)
        else:
            f.write(i[0].replace(',', '') + ',')  # 1st line write in csv
            f.write(i[1].replace(',', '') + ',')  # 2nd line write in csv
            f.write(i[2].replace(',', '').replace('/c/', '/pic/') + '\n')  # last line write in csv
    wb2.save(r'C:\Users\jayka\PycharmProjects\pythonProject\test\nodata.xlsx')  # nodata
    f.close()


#links.csv(for script 1) ---> links2.csv(for script 2)